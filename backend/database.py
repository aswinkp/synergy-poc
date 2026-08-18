from __future__ import annotations

import hashlib
import json
import sqlite3
from collections import Counter
from contextlib import contextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook

from .config import DATABASE_PATH
from .schema import COLUMNS, HEADCOUNT_COLUMNS


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


@contextmanager
def connect() -> Iterator[sqlite3.Connection]:
    DATABASE_PATH.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(DATABASE_PATH)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    try:
        yield connection
        connection.commit()
    finally:
        connection.close()


def _fingerprint(path: Path) -> str:
    stat = path.stat()
    raw = f"{path.resolve()}:{stat.st_size}:{stat.st_mtime_ns}".encode()
    return hashlib.sha256(raw).hexdigest()


def _create_application_tables(db: sqlite3.Connection) -> None:
    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS app_meta (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS users (
            id TEXT PRIMARY KEY,
            email TEXT NOT NULL UNIQUE COLLATE NOCASE,
            name TEXT NOT NULL,
            password_hash TEXT NOT NULL,
            is_active INTEGER NOT NULL DEFAULT 1 CHECK(is_active IN (0, 1)),
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS chats (
            id TEXT PRIMARY KEY,
            user_id TEXT REFERENCES users(id),
            title TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS messages (
            id TEXT PRIMARY KEY,
            chat_id TEXT NOT NULL REFERENCES chats(id) ON DELETE CASCADE,
            role TEXT NOT NULL CHECK(role IN ('user', 'assistant')),
            content TEXT NOT NULL,
            visualization TEXT,
            attachment TEXT,
            created_at TEXT NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_messages_chat ON messages(chat_id, created_at);
        """
    )
    message_columns = {
        row["name"] for row in db.execute("PRAGMA table_info(messages)").fetchall()
    }
    if "attachment" not in message_columns:
        db.execute("ALTER TABLE messages ADD COLUMN attachment TEXT")
    chat_columns = {
        row["name"] for row in db.execute("PRAGMA table_info(chats)").fetchall()
    }
    if "user_id" not in chat_columns:
        db.execute("ALTER TABLE chats ADD COLUMN user_id TEXT REFERENCES users(id)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_chats_user_updated ON chats(user_id, updated_at DESC)")


def _database_stats(db: sqlite3.Connection) -> dict[str, Any]:
    stats = db.execute(
        """
        SELECT
            (SELECT COUNT(*) FROM learning_records) AS records,
            (SELECT COUNT(DISTINCT employee_id) FROM learning_records) AS learning_employees,
            (SELECT COUNT(*) FROM employee_headcount) AS headcount_records,
            (SELECT COUNT(*) FROM employees WHERE is_in_headcount = 1) AS headcount_employees,
            (
                SELECT COUNT(DISTINCT l.employee_id)
                FROM learning_records AS l
                JOIN employees AS e ON e.employee_id = l.employee_id
                WHERE e.is_in_headcount = 1
            ) AS matched_learning_employees
        """
    ).fetchone()
    return dict(stats)


def initialize_database(workbook_path: Path, headcount_path: Path | None = None) -> dict[str, Any]:
    learning_fingerprint = _fingerprint(workbook_path)
    headcount_fingerprint = _fingerprint(headcount_path) if headcount_path else ""
    required_objects = {"learning_records", "employees", "employee_headcount", "employee_learning_summary"}

    with connect() as db:
        _create_application_tables(db)
        stored = {
            row["key"]: row["value"]
            for row in db.execute(
                "SELECT key, value FROM app_meta WHERE key IN ('learning_workbook_fingerprint', 'headcount_workbook_fingerprint')"
            ).fetchall()
        }
        existing_objects = {
            row["name"]
            for row in db.execute(
                "SELECT name FROM sqlite_master WHERE name IN ('learning_records', 'employees', 'employee_headcount', 'employee_learning_summary')"
            ).fetchall()
        }
        if (
            existing_objects == required_objects
            and stored.get("learning_workbook_fingerprint") == learning_fingerprint
            and stored.get("headcount_workbook_fingerprint", "") == headcount_fingerprint
        ):
            return {**_database_stats(db), "refreshed": False}

    stats = _import_workbooks(
        workbook_path,
        learning_fingerprint,
        headcount_path,
        headcount_fingerprint,
    )
    return {**stats, "refreshed": True}


def _clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if hasattr(value, "isoformat") and not isinstance(value, str):
        return value.isoformat()
    return value


def _normalize_employee_id(value: Any) -> str:
    return "".join(str(value or "").split()).upper()


def _read_workbook_rows(
    path: Path,
    columns: list[tuple[str, str, str]],
    *,
    header_row: int,
    first_data_row: int,
) -> list[tuple[Any, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        header_values = next(
            sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
        )
        actual_headers = [str(value).strip() if value is not None else "" for value in header_values[: len(columns)]]
        expected_headers = [label for _, label, _ in columns]
        headers_match = all(
            actual == expected
            or (actual == "Category" and expected in {"Learning Category", "Employee Category"})
            for actual, expected in zip(actual_headers, expected_headers)
        )
        if not headers_match:
            raise ValueError(
                f"{path.name} does not match the expected column layout. "
                f"Expected {expected_headers!r}, received {actual_headers!r}."
            )

        rows: list[tuple[Any, ...]] = []
        for row_number, raw in enumerate(
            sheet.iter_rows(min_row=first_data_row, values_only=True),
            start=first_data_row,
        ):
            if not any(value is not None and str(value).strip() for value in raw):
                continue
            values = [_clean_value(value) for value in raw[: len(columns)]]
            values.extend([None] * (len(columns) - len(values)))
            employee_id = _normalize_employee_id(values[0])
            if not employee_id:
                raise ValueError(f"{path.name} row {row_number} has no Employee ID.")
            values[0] = employee_id
            rows.append(tuple(values))
        return rows
    finally:
        workbook.close()


def _canonical_employees(
    learning_rows: list[tuple[Any, ...]],
    headcount_rows: list[tuple[Any, ...]],
) -> list[tuple[Any, ...]]:
    learning_indexes = {name: index for index, (name, _, _) in enumerate(COLUMNS)}
    headcount_indexes = {name: index for index, (name, _, _) in enumerate(HEADCOUNT_COLUMNS)}
    learning_by_employee: dict[str, tuple[Any, ...]] = {}
    headcount_by_employee: dict[str, tuple[Any, ...]] = {}
    headcount_counts = Counter(row[0] for row in headcount_rows)

    for row in learning_rows:
        learning_by_employee.setdefault(row[0], row)
    for row in headcount_rows:
        # The source contains exact duplicates and a few manager-only conflicts with no
        # effective-date distinction. Preserve every raw row and use the first export row
        # as the deterministic canonical profile.
        headcount_by_employee.setdefault(row[0], row)

    employee_rows: list[tuple[Any, ...]] = []
    for employee_id in sorted(set(learning_by_employee) | set(headcount_by_employee)):
        if employee_id in headcount_by_employee:
            profile = list(headcount_by_employee[employee_id])
        else:
            learning_row = learning_by_employee[employee_id]
            profile = [None] * len(HEADCOUNT_COLUMNS)
            profile[0] = employee_id
            for name, headcount_index in headcount_indexes.items():
                learning_index = learning_indexes.get(name)
                if learning_index is not None:
                    profile[headcount_index] = learning_row[learning_index]
        employee_rows.append(
            tuple(profile)
            + (
                headcount_counts.get(employee_id, 0),
                int(employee_id in headcount_by_employee),
                int(employee_id in learning_by_employee),
            )
        )
    return employee_rows


def _column_names(columns: list[tuple[str, str, str]]) -> str:
    return ", ".join(f'"{name}"' for name, _, _ in columns)


def _import_workbooks(
    learning_path: Path,
    learning_fingerprint: str,
    headcount_path: Path | None,
    headcount_fingerprint: str,
) -> dict[str, Any]:
    learning_rows = _read_workbook_rows(
        learning_path,
        COLUMNS,
        header_row=4,
        first_data_row=5,
    )
    headcount_rows = (
        _read_workbook_rows(
            headcount_path,
            HEADCOUNT_COLUMNS,
            header_row=1,
            first_data_row=2,
        )
        if headcount_path
        else []
    )
    employee_rows = _canonical_employees(learning_rows, headcount_rows)

    learning_names = _column_names(COLUMNS)
    learning_placeholders = ", ".join("?" for _ in COLUMNS)
    headcount_names = _column_names(HEADCOUNT_COLUMNS)
    headcount_placeholders = ", ".join("?" for _ in HEADCOUNT_COLUMNS)
    employee_extra_names = '"headcount_record_count", "is_in_headcount", "has_learning_records"'
    employee_placeholders = ", ".join("?" for _ in range(len(HEADCOUNT_COLUMNS) + 3))
    employee_definitions = ",\n".join(
        ['"employee_id" TEXT PRIMARY KEY']
        + [f'"{name}" {sql_type}' for name, _, sql_type in HEADCOUNT_COLUMNS[1:]]
        + [
            '"headcount_record_count" INTEGER NOT NULL',
            '"is_in_headcount" INTEGER NOT NULL CHECK(is_in_headcount IN (0, 1))',
            '"has_learning_records" INTEGER NOT NULL CHECK(has_learning_records IN (0, 1))',
        ]
    )
    learning_definitions = ",\n".join(
        [
            '"employee_id" TEXT NOT NULL REFERENCES employees(employee_id)',
            *[f'"{name}" {sql_type}' for name, _, sql_type in COLUMNS[1:]],
        ]
    )
    headcount_definitions = ",\n".join(
        [
            '"employee_id" TEXT NOT NULL REFERENCES employees(employee_id)',
            *[f'"{name}" {sql_type}' for name, _, sql_type in HEADCOUNT_COLUMNS[1:]],
        ]
    )

    with connect() as db:
        _create_application_tables(db)
        db.executescript(
            f"""
            DROP VIEW IF EXISTS employee_learning_summary;
            DROP TABLE IF EXISTS learning_records;
            DROP TABLE IF EXISTS employee_headcount;
            DROP TABLE IF EXISTS employees;

            CREATE TABLE employees ({employee_definitions});
            CREATE TABLE employee_headcount (
                headcount_row_id INTEGER PRIMARY KEY,
                {headcount_definitions}
            );
            CREATE TABLE learning_records (
                row_id INTEGER PRIMARY KEY,
                {learning_definitions}
            );
            """
        )
        db.executemany(
            f"INSERT INTO employees ({headcount_names}, {employee_extra_names}) VALUES ({employee_placeholders})",
            employee_rows,
        )
        if headcount_rows:
            db.executemany(
                f"INSERT INTO employee_headcount ({headcount_names}) VALUES ({headcount_placeholders})",
                headcount_rows,
            )
        db.executemany(
            f"INSERT INTO learning_records ({learning_names}) VALUES ({learning_placeholders})",
            learning_rows,
        )

        for column in ("employee_id", "status", "course_name", "company", "business_unit", "learning_category"):
            db.execute(f'CREATE INDEX idx_learning_{column} ON learning_records("{column}")')
        for column in ("company", "business_unit", "current_department", "job_level", "manager_id", "generation", "active_status"):
            db.execute(f'CREATE INDEX idx_employees_{column} ON employees("{column}")')
        db.execute("CREATE INDEX idx_headcount_employee_id ON employee_headcount(employee_id)")

        db.execute(
            """
            CREATE VIEW employee_learning_summary AS
            SELECT
                e.*,
                COUNT(l.row_id) AS learning_assignments,
                SUM(CASE WHEN l.status = 'Completed' THEN 1 ELSE 0 END) AS completed_assignments,
                SUM(CASE WHEN l.status = 'Not Started' THEN 1 ELSE 0 END) AS not_started_assignments,
                SUM(CASE WHEN l.status = 'In Progress' THEN 1 ELSE 0 END) AS in_progress_assignments,
                ROUND(
                    100.0 * SUM(CASE WHEN l.status = 'Completed' THEN 1 ELSE 0 END)
                    / NULLIF(COUNT(l.row_id), 0),
                    1
                ) AS completion_rate
            FROM employees AS e
            LEFT JOIN learning_records AS l ON l.employee_id = e.employee_id
            GROUP BY e.employee_id
            """
        )

        metadata = {
            "learning_workbook_fingerprint": learning_fingerprint,
            "learning_workbook_name": learning_path.name,
            "headcount_workbook_fingerprint": headcount_fingerprint,
            "headcount_workbook_name": headcount_path.name if headcount_path else "",
        }
        db.executemany(
            "INSERT INTO app_meta(key, value) VALUES(?, ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            metadata.items(),
        )
        return _database_stats(db)


def rows_as_dicts(rows: list[sqlite3.Row]) -> list[dict[str, Any]]:
    return [{key: row[key] for key in row.keys()} for row in rows]


def load_visualization(value: str | None) -> dict[str, Any] | None:
    return json.loads(value) if value else None


def load_attachment(value: str | None) -> dict[str, Any] | None:
    return json.loads(value) if value else None
